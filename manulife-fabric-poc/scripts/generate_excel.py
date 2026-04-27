"""Generate deal-sizing Excel workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
title_font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
currency_fmt = "$#,##0"
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap_align = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def write_table(ws, start_row, headers, data, col_widths=None, currency_cols=None):
    currency_cols = currency_cols or []
    for i, h in enumerate(headers):
        ws.cell(row=start_row, column=i + 1, value=h)
    style_header_row(ws, start_row, len(headers))
    for r, row_data in enumerate(data):
        for c, val in enumerate(row_data):
            cell = ws.cell(row=start_row + 1 + r, column=c + 1, value=val)
            cell.border = thin_border
            cell.alignment = wrap_align
            if c in currency_cols and isinstance(val, (int, float)):
                cell.number_format = currency_fmt
    if col_widths:
        for i, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w
    return start_row + 1 + len(data)


# === Sheet 1: Deal Summary ===
ws = wb.active
ws.title = "Deal Summary"
ws.cell(row=1, column=1, value="Manulife Fabric Opportunity - Deal Sizing (SWAG)").font = title_font
ws.merge_cells("A1:D1")
ws.cell(row=2, column=1, value="Internal Use Only | April 2026").font = Font(italic=True, color="666666")

row = 4
ws.cell(row=row, column=1, value="Deal Summary").font = section_font
row += 1
write_table(ws, row,
    ["Metric", "Low Estimate", "High Estimate", "Notes"],
    [
        ["POC Deal", 150000, 200000, "3-4 month engagement"],
        ["Year 1 Production", 1000000, 2000000, "Single BU deployment"],
        ["Annual Run Rate (Yr 2+)", 600000, 1500000, "Licensing + managed services"],
        ["3-Year TCV (Base Case)", 3000000, 6000000, "2-3 BUs, moderate expansion"],
        ["3-Year TCV (Upside)", 6000000, 12000000, "Enterprise-wide, Copilot at scale"],
    ],
    [30, 18, 18, 40], [1, 2])

# === Sheet 2: POC Phase ===
ws2 = wb.create_sheet("POC Phase")
ws2.cell(row=1, column=1, value="POC / Pilot Phase (3-4 Months)").font = section_font
write_table(ws2, 3,
    ["Component", "Monthly ($)", "Total 4mo ($)", "Notes"],
    [
        ["Fabric Capacity (F64)", 7000, 28000, "Minimum viable for POC"],
        ["Azure AI Search (Basic)", 1000, 4000, "Single index, low query volume"],
        ["Azure OpenAI (S0)", 2000, 8000, "Embeddings + GPT-4o, low throughput"],
        ["Power BI Pro (10 seats)", 1000, 4000, "Semantic model dev and testing"],
        ["Azure Subtotal", 11000, 44000, ""],
        ["Consulting - Arch & Impl", "", 112500, "Midpoint of $75K-$150K range"],
        ["", "", "", ""],
        ["POC TOTAL", "", 156500, "Target: $150K-$200K"],
    ],
    [35, 15, 15, 40], [1, 2])

# === Sheet 3: Year 1 Production ===
ws3 = wb.create_sheet("Year 1 Production")
ws3.cell(row=1, column=1, value="Year 1 - Production Deployment").font = section_font
write_table(ws3, 3,
    ["Component", "Low Annual ($)", "High Annual ($)", "Notes"],
    [
        ["Fabric Capacity (F128-F256)", 200000, 500000, "Production workloads"],
        ["Power BI Premium (P1-P2/PPU)", 60000, 250000, "Capacity vs seat model"],
        ["Azure OpenAI", 100000, 250000, "Production query volume"],
        ["Azure AI Search (Std S1-S2)", 50000, 120000, "Vector search, replicas"],
        ["Supporting Azure (KV/Storage)", 10000, 30000, "Infrastructure"],
        ["Azure + Licensing Subtotal", 420000, 1150000, ""],
        ["Implementation & Build-out", 300000, 800000, "Hardening, security, rollout"],
        ["", "", "", ""],
        ["YEAR 1 TOTAL", 720000, 1950000, "Target: $1M-$2M"],
    ],
    [38, 18, 18, 42], [1, 2])

# === Sheet 4: Steady State ===
ws4 = wb.create_sheet("Steady State (Yr 2+)")
ws4.cell(row=1, column=1, value="Steady-State Annual Run Rate").font = section_font
write_table(ws4, 3,
    ["Component", "Low Annual ($)", "High Annual ($)", "Notes"],
    [
        ["Fabric Capacity", 250000, 600000, "Grows with data volume"],
        ["Power BI Licensing", 60000, 250000, "Stable unless seats expand"],
        ["Azure AI Services", 150000, 350000, "Scales with query volume"],
        ["Managed Services / Support", 150000, 400000, "L2/L3, tuning, enhancements"],
        ["", "", "", ""],
        ["ANNUAL RUN RATE", 610000, 1600000, ""],
    ],
    [38, 18, 18, 42], [1, 2])

# === Sheet 5: 3-Year Scenarios ===
ws5 = wb.create_sheet("3-Year Scenarios")
ws5.cell(row=1, column=1, value="3-Year Total Contract Value Scenarios").font = section_font
row = write_table(ws5, 3,
    ["Scenario", "Assumptions", "Low TCV ($)", "High TCV ($)"],
    [
        ["Conservative", "Single BU, contained scope, no major expansion", 1500000, 3000000],
        ["Base Case", "2-3 BUs adopt, moderate seat expansion", 3000000, 6000000],
        ["Upside", "Enterprise-wide, Copilot at scale, multiple AI use cases", 6000000, 12000000],
    ],
    [18, 50, 18, 18], [2, 3])

row += 2
ws5.cell(row=row, column=1, value="Expansion Vectors").font = section_font
row += 1
write_table(ws5, row,
    ["Vector", "Description", "Revenue Multiplier"],
    [
        ["Additional BUs", "Group Benefits, Individual Insurance, Wealth, Retirement", "2x-4x"],
        ["Data Domains", "Actuarial, risk, compliance, finance, operations", "1.5x-2x"],
        ["Copilot Seats", "Pilot (50-100) to enterprise (1,000-5,000+)", "Significant uplift"],
        ["Data Volume", "More sources, history, real-time streams", "1.3x-2x"],
        ["Geography", "US and Asia operations", "1.5x-2x"],
        ["Advanced AI", "Agentic workflows, auto triage, advisor copilot", "Incremental"],
    ],
    [25, 55, 25])

# === Sheet 6: Risks ===
ws6 = wb.create_sheet("Risks & Assumptions")
ws6.cell(row=1, column=1, value="Risks to Deal Size").font = section_font
row = write_table(ws6, 3,
    ["Risk", "Impact", "Likelihood", "Mitigation"],
    [
        ["Existing Microsoft ELA covers Fabric", "Reduces net-new licensing", "Medium", "Validate licensing early"],
        ["Data Agent stays in preview", "Delays production commitment", "Medium", "Pivot to PBI Copilot"],
        ["Competing platform entrenched", "Limits to specific BUs", "Low-Medium", "Position as complementary"],
        ["Budget freeze / procurement delay", "Extends sales cycle", "Low", "Align to budget cycles"],
        ["Internal build preference", "Reduces consulting scope", "Medium", "Demonstrate POC velocity"],
    ],
    [32, 28, 15, 38])

row += 2
ws6.cell(row=row, column=1, value="Key Assumptions").font = section_font
row += 1
for a in [
    "Manulife does not already have significant Fabric / Power BI Premium licensing",
    "Fabric Data Agent reaches GA in a compatible timeframe",
    "POC successfully demonstrates value to business stakeholders",
    "Partner-delivered consulting is the preferred delivery model",
    "No competing platform commitment blocks Fabric adoption",
]:
    ws6.cell(row=row, column=1, value="  " + a).alignment = wrap_align
    ws6.column_dimensions["A"].width = 80
    row += 1

# === Sheet 7: Next Steps ===
ws7 = wb.create_sheet("Next Steps")
ws7.cell(row=1, column=1, value="Recommended Next Steps").font = section_font
write_table(ws7, 3,
    ["#", "Action", "Owner", "Timeline"],
    [
        [1, "Validate current Microsoft licensing posture (ELA, Fabric, PBI)", "Account Team", "Week 1"],
        [2, "Identify executive sponsor and budget holder", "Account Team", "Week 1-2"],
        [3, "Present POC results and reference architecture", "Delivery Team", "Week 2"],
        [4, "Scope production pilot (single BU, defined user group)", "Joint", "Week 3-4"],
        [5, "Develop SOW for production phase", "Delivery Team", "Week 4-5"],
        [6, "Engage Microsoft co-sell / FastTrack team", "Account Team", "Week 2-3"],
        [7, "Align on success criteria and go/no-go for expansion", "Joint", "Month 3-4"],
    ],
    [5, 55, 18, 15])

# === Sheet 8: Competitive ===
ws8 = wb.create_sheet("Competitive")
ws8.cell(row=1, column=1, value="Competitive Landscape").font = section_font
write_table(ws8, 3,
    ["Competitor", "Threat Level", "Our Positioning"],
    [
        ["Databricks + Unity Catalog", "Medium-High", "Fabric: unified platform, no BI/AI stitching"],
        ["Snowflake + Cortex", "Medium", "Power BI + Copilot integration differentiated"],
        ["AWS (Bedrock + Redshift)", "Low-Medium", "Manulife likely has Azure affinity"],
        ["Internal / Custom Build", "Medium", "POC demonstrates speed-to-value"],
    ],
    [28, 15, 55])

filepath = "C:/Users/anuragdhuria/OneDrive - Microsoft/Documents/GitHub/desktop-tutorial/manulife-fabric-poc/docs/deal-sizing-swag.xlsx"
wb.save(filepath)
print(f"Excel saved: {filepath}")
print(f"Sheets: {wb.sheetnames}")
